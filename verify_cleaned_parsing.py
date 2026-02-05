#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""验证HTML实体清理后的分解结果"""
import openpyxl

excel_file = "/Users/yuanyuqing/Documents/code/schoolProject/backend/outputs/协议_20260205172324.xlsx"

wb = openpyxl.load_workbook(excel_file)
ws = wb.active

print(f"\n{'='*100}")
print(f"HTML实体清理后的分解验证")
print(f"{'='*100}\n")

# 获取列标题
headers = []
for col_idx in range(1, ws.max_column + 1):
    header = ws.cell(1, col_idx).value
    headers.append(str(header) if header else "")

# 查找信源、信宿相关的列
src_mach_idx = None
dst_mach_idx = None

for idx, h in enumerate(headers):
    if '信源机器码' in str(h):
        src_mach_idx = idx + 1
    if '信宿机器码' in str(h):
        dst_mach_idx = idx + 1

print(f"【检查结果1表】(第2行):")
print(f"  原始值应该是: BC → RT1-SA0-模式码0x03")
print(f"  提取结果:")
row_idx = 2
if src_mach_idx:
    src = ws.cell(row_idx, src_mach_idx).value
    print(f"    信源机器码: {src} {'✓' if src and 'BC' in str(src) and 'è' not in str(src) else '✗'}")
if dst_mach_idx:
    dst = ws.cell(row_idx, dst_mach_idx).value
    print(f"    信宿机器码: {dst} {'✓' if dst and 'RT1-SA0' in str(dst) else '✗'}")

print(f"\n【检查结果2表】(第3行):")
print(f"  原始值应该是: BC → RT1-SA0-模式码0x04")
print(f"  提取结果:")
row_idx = 3
if src_mach_idx:
    src = ws.cell(row_idx, src_mach_idx).value
    print(f"    信源机器码: {src} {'✓' if src and 'BC' in str(src) and 'è' not in str(src) else '✗'}")
if dst_mach_idx:
    dst = ws.cell(row_idx, dst_mach_idx).value
    print(f"    信宿机器码: {dst} {'✓' if dst and 'RT1-SA0' in str(dst) else '✗'}")

wb.close()
