# -*- coding: utf-8 -*-
"""
Offline acceptance check for Win7 packaging.

Run on the Windows/Python 3.8 build machine before build_exe.bat:
    python offline_acceptance_check.py "C:\\path\\to\\测试协议20260508.docx"
"""

import argparse
import os
import sys


ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)


def _load_workbook_rows(path):
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [cell.value or '' for cell in ws[1]]
    rows = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        row = {}
        has_value = False
        for idx, value in enumerate(excel_row):
            if idx >= len(headers) or not headers[idx]:
                continue
            if value not in (None, ''):
                has_value = True
            row[headers[idx]] = value
        if has_value:
            rows.append(row)
    return headers, rows


def _row_text(row):
    return ' '.join(str(v) for v in row.values() if v not in (None, ''))


def run_check(docx_path, output_dir):
    from backend.services.table_detector import DocumentParser
    from backend.routes.extract import build_processed_tables
    from backend.services.excel_exporter import ExcelExporter
    from backend.services.output_postprocessor import has_effective_data, row_contains_text

    parser = DocumentParser()
    parsed = parser.parse(docx_path)
    processed = build_processed_tables(
        parsed.get('tables', []),
        output_options={'remove_crc_checksum': True},
    )
    trailing_crc_tables = []
    for table in processed:
        effective_rows = [row for row in table.get('data_rows', []) if has_effective_data(row)]
        if effective_rows and row_contains_text(effective_rows[-1], 'CRC校验'):
            trailing_crc_tables.append(table.get('msg_name', ''))

    os.makedirs(output_dir, exist_ok=True)
    output_path = ExcelExporter(output_dir).export_with_template(processed, 'acceptance')
    headers, rows = _load_workbook_rows(output_path)

    note_rows = [idx + 2 for idx, row in enumerate(rows) if _row_text(row).startswith(('注：', '注:'))]
    id_count = sum(1 for row in rows if row.get('ID') not in (None, '', '—', '-'))
    bit_rows = [
        row for row in rows
        if row.get('子内容') not in (None, '', '—', '-')
        and row.get('类型（bit）') not in (None, '', '—', '-')
    ]
    bit_content_duplicates = [
        row for row in bit_rows
        if row.get('内容') not in (None, '', '—', '-')
        and str(row.get('内容')).strip() == str(row.get('子内容')).strip()
    ]

    print('Output:', output_path)
    print('Rows:', len(rows))
    print('ID rows:', id_count)
    print('Bit child rows:', len(bit_rows))
    print('Note rows:', note_rows)
    print('Trailing CRC tables:', trailing_crc_tables)
    print('Bit content duplicates:', len(bit_content_duplicates))

    failures = []
    if trailing_crc_tables:
        failures.append('some tables still end with CRC校验')
    if note_rows:
        failures.append('output still contains 注：/注: rows')
    if id_count == 0:
        failures.append('ID column has no recognized value')
    if bit_rows and bit_content_duplicates:
        failures.append('bit child rows duplicate 子内容 into 内容')

    if failures:
        print('\nFAILED:')
        for item in failures:
            print('-', item)
        return 1

    print('\nPASSED')
    return 0


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('docx_path')
    parser.add_argument('--output-dir', default=os.path.join(ROOT_DIR, 'backend', 'outputs'))
    args = parser.parse_args()

    if not os.path.exists(args.docx_path):
        print('Input file not found:', args.docx_path)
        return 2

    return run_check(args.docx_path, args.output_dir)


if __name__ == '__main__':
    raise SystemExit(main())
