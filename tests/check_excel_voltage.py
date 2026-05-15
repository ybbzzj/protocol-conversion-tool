#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检查最新生成的 Excel 文件中的电压字段内容
"""
import sys
import os
import glob

sys.path.insert(0, 'backend')

from openpyxl import load_workbook

# 找到最新的 Excel 文件
files = sorted(glob.glob('tests/backend/outputs/*.xlsx'))
if not files:
    print("未找到 Excel 文件")
    sys.exit(1)

latest_file = files[-1]
print(f"检查最新文件：{os.path.basename(latest_file)}\n")

wb = load_workbook(latest_file)
ws = wb.active

# 读取表头
headers = [cell.value for cell in ws[1]]
print(f"表头：{headers}\n")

print(f"共有 {ws.max_row - 1} 行数据\n")

# 显示所有行
print("所有数据行:")
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
    print(f"\n行 {row_idx}:")
    for col_idx, value in enumerate(row):
        if value is not None and str(value).strip():
            header_name = headers[col_idx] if col_idx < len(headers) else f"列{col_idx+1}"
            print(f"  {header_name}: {value}")
