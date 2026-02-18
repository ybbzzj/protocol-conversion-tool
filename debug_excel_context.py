#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'backend'))

from backend.app import create_app

def test_excel_export_with_context():
    """在Flask应用上下文中测试Excel导出"""
    print("🧪 Flask上下文中的Excel导出测试")
    print("=" * 50)
    
    app = create_app()
    
    with app.app_context():
        from backend.services.excel_exporter import ExcelExporter
        
        # 准备测试数据
        test_tables_data = [{
            'msg_name': '测试协议表',
            'data_rows': [
                {
                    'ID': '1',
                    '内容': '时间',
                    '转换类型': 'UINT32',
                    '类型（bit）': '32',
                    '判读公式': '0~4294967295',
                    '单位': 'ms',
                    '备注': '测试数据'
                },
                {
                    'ID': '2', 
                    '内容': '电压',
                    '转换类型': 'FLOAT',
                    '类型（bit）': '32',
                    '判读公式': '0~5.0',
                    '单位': 'V',
                    '备注': '电压测量'
                }
            ],
            'meta': {
                '信息名称': '测试协议',
                '信源': '测试设备',
                '信宿': '目标系统'
            }
        }]
        
        print("✅ 测试数据准备完成")
        
        # 测试Excel导出
        try:
            exporter = ExcelExporter('backend/outputs')
            output_file = exporter.export_with_template(test_tables_data, 'context_test')
            print(f"✅ Excel导出成功: {output_file}")
            
            # 验证生成的文件
            print("\n🔍 验证生成文件:")
            if os.path.exists(output_file):
                # 检查文件类型
                import subprocess
                result = subprocess.run(['file', output_file], capture_output=True, text=True)
                print(f"  文件类型: {result.stdout.strip()}")
                
                # 检查文件大小
                size = os.path.getsize(output_file)
                print(f"  文件大小: {size} bytes")
                
                # 尝试打开文件
                from openpyxl import load_workbook
                try:
                    wb = load_workbook(output_file)
                    ws = wb.active
                    print(f"  ✅ 可以用openpyxl打开")
                    print(f"  工作表: {wb.sheetnames}")
                    print(f"  行数: {ws.max_row}, 列数: {ws.max_column}")
                    
                    # 显示前几行内容
                    print("  前5行内容:")
                    for row in range(1, min(6, ws.max_row + 1)):
                        row_data = []
                        for col in range(1, min(8, ws.max_column + 1)):
                            cell_value = ws.cell(row=row, column=col).value
                            row_data.append(str(cell_value) if cell_value is not None else "")
                        print(f"    行{row}: {row_data}")
                        
                except Exception as e:
                    print(f"  ❌ openpyxl打开失败: {e}")
                    
        except Exception as e:
            print(f"❌ 导出过程出错: {e}")
            import traceback
            traceback.print_exc()

def create_alternative_exporter():
    """创建不依赖Flask上下文的简化导出器"""
    print("\n" + "=" * 50)
    print("🔧 创建简化导出器（无Flask依赖）")
    
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
    import shutil
    
    def simple_export(tables_data, output_path):
        """简化版Excel导出"""
        # 复制模板
        template_path = os.path.join(os.getcwd(), 'word', 'csvfile', '协议模板.xlsx')
        shutil.copy(template_path, output_path)
        
        # 加载工作簿
        wb = load_workbook(output_path)
        ws = wb.active
        
        # 获取模板表头
        template_headers = [cell.value if cell.value else "" for cell in ws[1]]
        current_row = 2
        
        # 填充数据
        for table in tables_data:
            # 填充元数据行
            meta = table.get('meta', {})
            if meta:
                for key, value in meta.items():
                    if value and str(value).strip():
                        ws.cell(row=current_row, column=1, value=f"{key}: {value}")
                        current_row += 1
            
            # 填充数据行
            for data_row in table['data_rows']:
                for col_idx, col_name in enumerate(template_headers, 1):
                    if col_name and col_name in data_row:
                        cell_value = data_row[col_name]
                        if cell_value is not None:
                            ws.cell(row=current_row, column=col_idx, value=str(cell_value))
                current_row += 1
        
        wb.save(output_path)
        return output_path
    
    # 测试简化导出器
    test_data = [{
        'msg_name': '简化测试表',
        'data_rows': [
            {'ID': '1', '内容': '测试时间', '转换类型': 'UINT32', '单位': 'ms'},
            {'ID': '2', '内容': '测试电压', '转换类型': 'FLOAT', '单位': 'V'}
        ],
        'meta': {'信息名称': '简化测试协议'}
    }]
    
    output_file = 'simple_export_no_flask.xlsx'
    try:
        result = simple_export(test_data, output_file)
        print(f"✅ 简化导出成功: {result}")
        
        # 验证文件
        import subprocess
        result = subprocess.run(['file', output_file], capture_output=True, text=True)
        print(f"  文件类型: {result.stdout.strip()}")
        print(f"  文件大小: {os.path.getsize(output_file)} bytes")
        
        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        ws = wb.active
        print(f"  ✅ 可以打开，{ws.max_row}行 x {ws.max_column}列")
        
    except Exception as e:
        print(f"❌ 简化导出失败: {e}")

if __name__ == "__main__":
    test_excel_export_with_context()
    create_alternative_exporter()
    print("\n" + "=" * 50)
    print("📋 测试完成")