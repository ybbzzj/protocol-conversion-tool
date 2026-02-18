#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'backend'))

from backend.services.excel_exporter import ExcelExporter
from backend.services.table_detector import DocumentParser
from backend.services.data_cleaner import DataProcessor
from backend.services.field_matcher import FieldMatcher

def test_full_export_process():
    """测试完整的Excel导出流程"""
    print("🧪 完整导出流程测试")
    print("=" * 50)
    
    # 1. 准备测试数据
    test_tables_data = [{
        'msg_name': '测试协议表',
        'data_rows': [
            {
                'ID': '1',
                '内容': '时间',
                '转换类型': 'UINT32',
                '类型（bit）': '32',
                '判读公式': '0~4294967295',
                '单位': 'ms',
                '备注': '测试数据'
            },
            {
                'ID': '2', 
                '内容': '电压',
                '转换类型': 'FLOAT',
                '类型（bit）': '32',
                '判读公式': '0~5.0',
                '单位': 'V',
                '备注': '电压测量'
            }
        ],
        'meta': {
            '信息名称': '测试协议',
            '信源': '测试设备',
            '信宿': '目标系统'
        }
    }]
    
    print("✅ 测试数据准备完成")
    
    # 2. 测试Excel导出
    try:
        exporter = ExcelExporter('backend/outputs')
        output_file = exporter.export_with_template(test_tables_data, 'debug_test')
        print(f"✅ Excel导出成功: {output_file}")
        
        # 3. 验证生成的文件
        print("\n🔍 验证生成文件:")
        if os.path.exists(output_file):
            # 检查文件类型
            import subprocess
            result = subprocess.run(['file', output_file], capture_output=True, text=True)
            print(f"  文件类型: {result.stdout.strip()}")
            
            # 检查文件大小
            size = os.path.getsize(output_file)
            print(f"  文件大小: {size} bytes")
            
            # 尝试打开文件
            from openpyxl import load_workbook
            try:
                wb = load_workbook(output_file)
                ws = wb.active
                print(f"  ✅ 可以用openpyxl打开")
                print(f"  工作表: {wb.sheetnames}")
                print(f"  行数: {ws.max_row}, 列数: {ws.max_column}")
                
                # 显示前几行内容
                print("  前3行内容:")
                for row in range(1, min(4, ws.max_row + 1)):
                    row_data = []
                    for col in range(1, min(6, ws.max_column + 1)):
                        cell_value = ws.cell(row=row, column=col).value
                        row_data.append(str(cell_value) if cell_value is not None else "")
                    print(f"    行{row}: {row_data}")
                    
            except Exception as e:
                print(f"  ❌ openpyxl打开失败: {e}")
                # 尝试用zip检查文件结构
                import zipfile
                try:
                    with zipfile.ZipFile(output_file, 'r') as zf:
                        print("  ZIP文件结构正常")
                        print("  包含文件:", zf.namelist())
                except Exception as ze:
                    print(f"  ❌ ZIP结构也损坏: {ze}")
        else:
            print("❌ 输出文件不存在")
            
    except Exception as e:
        print(f"❌ 导出过程出错: {e}")
        import traceback
        traceback.print_exc()

def compare_with_simple_export():
    """对比简单导出和模板导出"""
    print("\n" + "=" * 50)
    print("📊 简单导出 vs 模板导出对比")
    
    # 简单导出（不使用模板）
    from openpyxl import Workbook
    wb_simple = Workbook()
    ws_simple = wb_simple.active
    ws_simple.title = "简单测试"
    
    headers = ['ID', '内容', '类型', '值域', '单位']
    ws_simple.append(headers)
    
    data = [
        ['1', '时间', 'UINT32', '0~4294967295', 'ms'],
        ['2', '电压', 'FLOAT', '0~5.0', 'V']
    ]
    
    for row in data:
        ws_simple.append(row)
    
    simple_file = 'simple_export.xlsx'
    wb_simple.save(simple_file)
    print(f"✅ 简单导出完成: {simple_file}")
    
    # 验证两个文件
    files_to_check = ['simple_export.xlsx']
    if os.path.exists('backend/outputs/协议_debug_test.xlsx'):
        files_to_check.append('backend/outputs/协议_debug_test.xlsx')
    
    for filename in files_to_check:
        if os.path.exists(filename):
            print(f"\n🔍 检查 {filename}:")
            import subprocess
            result = subprocess.run(['file', filename], capture_output=True, text=True)
            print(f"  文件类型: {result.stdout.strip()}")
            print(f"  文件大小: {os.path.getsize(filename)} bytes")

if __name__ == "__main__":
    test_full_export_process()
    compare_with_simple_export()
    print("\n" + "=" * 50)
    print("📋 测试完成")