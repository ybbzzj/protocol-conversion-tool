#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
深度分析：源文档表格结构 + Excel输出结构
目标：彻底搞清楚每个表的结构、字段、合并单元格等细节
"""
from docx import Document
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from openpyxl import load_workbook
from openpyxl.styles import Font
import re

DOCX_PATH = '/Users/yuanyuqing/Documents/code/schoolProject/word/测试协议20260227.docx'
XLSX_PATH = '/Users/yuanyuqing/Documents/code/schoolProject/word/csvfile/转换结果20260227.xlsx'

# ==================== 分析DOCX ====================
doc = Document(DOCX_PATH)

# 提取文档中的所有元素（段落+表格），保留顺序和段落文字
elements = []
for el in doc.element.body:
    if isinstance(el, CT_P):
        text = el.text_content() if hasattr(el, 'text_content') else ''
        # 兜底：直接拼接所有<w:t>
        ns = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'
        t_nodes = el.findall(f'.//{{{ns}}}t')
        text = ''.join((t.text or '') for t in t_nodes).strip()
        if text:
            elements.append({'type': 'paragraph', 'text': text})
    elif isinstance(el, CT_Tbl):
        table = Table(el, doc)
        # 获取合并单元格信息
        rows_data = []
        for row_idx, row in enumerate(table.rows):
            cells = []
            for col_idx, cell in enumerate(row.cells):
                # 获取单元格文本
                text = cell.text.strip()
                # 检查合并信息（通过XML属性）
                tc = cell._tc
                # gridSpan: 水平合并
                gridSpan = 1
                gs = tc.find('.//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}gridSpan')
                if gs is not None:
                    val = gs.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val')
                    if val:
                        gridSpan = int(val)
                # vMerge: 垂直合并
                vMerge = None
                vm = tc.find('.//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}vMerge')
                if vm is not None:
                    v = vm.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val')
                    vMerge = v if v else 'continue'
                cells.append({
                    'text': text,
                    'col': col_idx,
                    'gridSpan': gridSpan,
                    'vMerge': vMerge
                })
            rows_data.append(cells)
        elements.append({'type': 'table', 'rows': rows_data, 'row_count': len(table.rows), 'col_count': len(table.columns)})

# 打印所有元素
print("="*140)
print("【文档元素顺序】（段落+表格）")
print("="*140)

table_counter = 0
for elem_idx, elem in enumerate(elements):
    if elem['type'] == 'paragraph':
        print(f"\n[段落] {elem['text']}")
    elif elem['type'] == 'table':
        table_counter += 1
        print(f"\n{'='*120}")
        print(f"[表格 #{table_counter}] 维度: {elem['row_count']}行 × {elem['col_count']}列")
        print(f"{'='*120}")
        for row_idx, row in enumerate(elem['rows']):
            row_display = []
            for cell in row:
                span_info = f"(span={cell['gridSpan']})" if cell['gridSpan'] > 1 else ""
                merge_info = f"(vMerge={cell['vMerge']})" if cell['vMerge'] else ""
                text = cell['text'][:35] if len(cell['text']) > 35 else cell['text']
                row_display.append(f"'{text}'{span_info}{merge_info}")
            print(f"  行{row_idx}: {row_display}")

print(f"\n\n总表格数: {table_counter}")

# ==================== 分析XLSX ====================
print("\n\n" + "="*140)
print("【Excel详细分析】")
print("="*140)

wb = load_workbook(XLSX_PATH)
ws = wb.active

# 获取合并单元格信息
print("\n【合并单元格】")
for merge in ws.merged_cells.ranges:
    print(f"  合并区域: {merge}")

# 获取列头
headers = [cell.value for cell in ws[1]]
print(f"\n【列名】({len(headers)}列)")
for i, h in enumerate(headers):
    print(f"  {i+1}: {h}")

# 逐行读取所有数据，包含字体颜色
print(f"\n【完整数据（含颜色信息）】")
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
    row_info = []
    for col_idx, cell in enumerate(row):
        val = cell.value
        if val is None:
            continue
        color = "black"
        if cell.font and cell.font.color:
            rgb = cell.font.color.rgb if hasattr(cell.font.color, 'rgb') else None
            if rgb and rgb not in ('00000000', 'FF000000', None, 'None'):
                color = f"rgb={rgb}"
        row_info.append(f"col{col_idx+1}[{headers[col_idx]}]='{val}'({color})")
    if row_info:
        print(f"\n行{row_idx}: {row_info}")
    else:
        print(f"\n行{row_idx}: (空行)")
