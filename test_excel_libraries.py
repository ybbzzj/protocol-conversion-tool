#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
from openpyxl import Workbook as OpenpyxlWorkbook
from openpyxl.styles import Font, Alignment
import xlsxwriter

def test_openpyxl():
    """使用openpyxl生成测试文件"""
    wb = OpenpyxlWorkbook()
    ws = wb.active
    ws.title = "测试表格"
    
    # 添加测试数据
    headers = ['ID', '名称', '类型', '值域', '单位']
    ws.append(headers)
    
    # 设置表头样式
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 添加数据行
    data_rows = [
        ['1', '时间', 'UINT32', '0~4294967295', 'ms'],
        ['2', '电压', 'FLOAT', '0~5.0', 'V'],
        ['3', '电流', 'FLOAT', '0~20.0', 'mA']
    ]
    
    for row_data in data_rows:
        ws.append(row_data)
    
    # 保存文件
    filename = 'test_openpyxl.xlsx'
    wb.save(filename)
    print(f"✅ openpyxl生成文件: {filename}")
    return filename

def test_xlsxwriter():
    """使用xlsxwriter生成测试文件"""
    filename = 'test_xlsxwriter.xlsx'
    workbook = xlsxwriter.Workbook(filename)
    worksheet = workbook.add_worksheet('测试表格')
    
    # 定义格式
    header_format = workbook.add_format({
        'bold': True,
        'align': 'center',
        'valign': 'vcenter'
    })
    
    # 添加表头
    headers = ['ID', '名称', '类型', '值域', '单位']
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)
    
    # 添加数据
    data_rows = [
        ['1', '时间', 'UINT32', '0~4294967295', 'ms'],
        ['2', '电压', 'FLOAT', '0~5.0', 'V'],
        ['3', '电流', 'FLOAT', '0~20.0', 'mA']
    ]
    
    for row, row_data in enumerate(data_rows, 1):
        for col, cell_data in enumerate(row_data):
            worksheet.write(row, col, cell_data)
    
    workbook.close()
    print(f"✅ xlsxwriter生成文件: {filename}")
    return filename

def validate_files(*filenames):
    """验证生成的文件"""
    import subprocess
    
    for filename in filenames:
        if not os.path.exists(filename):
            print(f"❌ 文件不存在: {filename}")
            continue
            
        print(f"\n🔍 验证文件: {filename}")
        
        # 检查文件类型
        result = subprocess.run(['file', filename], capture_output=True, text=True)
        print(f"  文件类型: {result.stdout.strip()}")
        
        # 检查文件大小
        size = os.path.getsize(filename)
        print(f"  文件大小: {size} bytes")
        
        # 尝试用openpyxl打开
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filename)
            ws = wb.active
            print(f"  ✅ openpyxl可以打开")
            print(f"  工作表: {wb.sheetnames}")
            print(f"  行数: {ws.max_row}, 列数: {ws.max_column}")
        except Exception as e:
            print(f"  ❌ openpyxl打开失败: {e}")

def main():
    print("🧪 Excel库对比测试")
    print("=" * 50)
    
    # 生成测试文件
    openpyxl_file = test_openpyxl()
    xlsxwriter_file = test_xlsxwriter()
    
    # 验证文件
    validate_files(openpyxl_file, xlsxwriter_file)
    
    print("\n" + "=" * 50)
    print("📋 测试完成")

if __name__ == "__main__":
    main()