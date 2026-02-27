#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
深度分析Part3: 精确分析每一个源表的具体内容和转换对应
重点：
1. 逐表打印完整内容（含合并单元格解析后实际值）
2. 分析每个表的"表头结构"（哪些列合并了，表头是几行）
3. 对比源表和Excel输出，找出精确映射规则
4. 研究不被提取的表格的共同特征
"""
from docx import Document
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from openpyxl import load_workbook

DOCX_PATH = '/Users/yuanyuqing/Documents/code/schoolProject/word/测试协议20260227.docx'
XLSX_PATH = '/Users/yuanyuqing/Documents/code/schoolProject/word/csvfile/转换结果20260227.xlsx'

doc = Document(DOCX_PATH)
ns = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'

# 构建元素列表
elements = []
for el in doc.element.body:
    if isinstance(el, CT_P):
        t_nodes = el.findall(f'.//{{{ns}}}t')
        text = ''.join((t.text or '') for t in t_nodes).strip()
        if text:
            elements.append({'type': 'para', 'text': text})
    elif isinstance(el, CT_Tbl):
        table = Table(el, doc)
        rows_data = []
        for row in table.rows:
            row_data = []
            for cell in row.cells:
                tc = cell._tc
                text = cell.text.strip()
                gs_node = tc.find(f'.//{{{ns}}}gridSpan')
                gridSpan = int(gs_node.get(f'{{{ns}}}val')) if gs_node is not None and gs_node.get(f'{{{ns}}}val') else 1
                vm_node = tc.find(f'.//{{{ns}}}vMerge')
                if vm_node is not None:
                    v = vm_node.get(f'{{{ns}}}val')
                    vMerge = v if v else 'continue'
                else:
                    vMerge = None
                row_data.append({'text': text, 'gridSpan': gridSpan, 'vMerge': vMerge})
            rows_data.append(row_data)
        elements.append({
            'type': 'table',
            'rows': rows_data,
            'row_count': len(table.rows),
            'col_count': len(table.columns)
        })

# 关联上下文段落
structured = []
last_paras = []
for elem in elements:
    if elem['type'] == 'para':
        last_paras.append(elem['text'])
    elif elem['type'] == 'table':
        structured.append({'paras': list(last_paras), 'table': elem})
        last_paras = []

# 解析合并单元格后的"逻辑视图"（把合并单元格展开）
def expand_table(rows_data):
    """将合并单元格展开为逻辑矩阵，方便分析"""
    if not rows_data:
        return []
    max_cols = max(len(row) for row in rows_data)
    # 构建一个 row x col 的矩阵，初始为 None
    result = [[None] * (max_cols * 2) for _ in range(len(rows_data))]
    vertical_fill = {}  # col -> (text, remaining)
    
    for row_idx, row in enumerate(rows_data):
        col_pos = 0
        for cell in row:
            # 跳过已被竖合并填充的列
            while col_pos in vertical_fill and vertical_fill[col_pos][1] > 0:
                result[row_idx][col_pos] = vertical_fill[col_pos][0]
                vertical_fill[col_pos] = (vertical_fill[col_pos][0], vertical_fill[col_pos][1] - 1)
                col_pos += 1
            
            text = cell['text']
            span = cell['gridSpan']
            
            if cell['vMerge'] == 'restart':
                # 记录竖向合并起点（但我们不知道延伸多少行，先放着）
                pass
            
            for s in range(span):
                if col_pos + s < len(result[row_idx]):
                    result[row_idx][col_pos + s] = text
            col_pos += span
    
    # 裁剪右侧None
    actual_cols = max_cols
    trimmed = []
    for row in result:
        trimmed.append([v if v is not None else '' for v in row[:actual_cols]])
    return trimmed

# ===== 逐表分析 =====
print("="*140)
print("【逐表详细分析】")
print("="*140)

for idx, item in enumerate(structured):
    t = item['table']
    paras = item['paras']
    rows = t['rows']
    
    print(f"\n{'='*140}")
    print(f"【源表 #{idx+1}】 {paras[-1][:80] if paras else '（无前置标题）'}")
    print(f"  维度: {t['row_count']}行 × {t['col_count']}列")
    print(f"  前置段落: {[p[:50] for p in paras[-3:]]}")
    
    # 逐行打印，显示合并信息
    print(f"\n  原始数据（含合并标注）:")
    for row_idx, row in enumerate(rows):
        row_display = []
        for c in row:
            text = c['text'][:25] if c['text'] else '""'
            extra = []
            if c['gridSpan'] > 1:
                extra.append(f"→{c['gridSpan']}列")
            if c['vMerge']:
                extra.append(f"↕{c['vMerge']}")
            cell_str = f"{text}{'[' + ','.join(extra) + ']' if extra else ''}"
            row_display.append(cell_str)
        print(f"  行{row_idx:2d}: {row_display}")

# ===== 关键对比：某设备装置测量数据3的字段表 vs Excel =====
print("\n\n" + "="*140)
print("【关键映射对比：源表字段列 → Excel列】")
print("="*140)

# 打印几个典型的字段定义表
target_tables = [5, 6, 7, 8, 9, 10, 11]  # 某设备装置测量数据3-5等
for idx in target_tables:
    if idx < len(structured):
        item = structured[idx]
        t = item['table']
        paras = item['paras']
        print(f"\n源表#{idx+1}: {paras[-1][:60] if paras else '无标题'}")
        for row_idx, row in enumerate(t['rows']):
            print(f"  行{row_idx}: {[c['text'][:20] for c in row]}")

# ===== Excel完整数据 =====
print("\n\n" + "="*140)
print("【Excel完整逐行分析】")
print("="*140)

wb = load_workbook(XLSX_PATH)
ws = wb.active
headers = [cell.value for cell in ws[1]]
print(f"表头: {headers}\n")

all_excel_rows = []
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
    row_dict = {}
    for col_idx, cell in enumerate(row):
        h = headers[col_idx]
        if h is None:
            continue
        val = cell.value
        color = "默认"
        if cell.font and cell.font.color:
            try:
                t_type = cell.font.color.type
                if t_type == 'rgb':
                    rgb = cell.font.color.rgb
                    if rgb and rgb not in ('00000000', 'FF000000'):
                        color = f"RGB:{rgb}"
                elif t_type == 'theme':
                    color = f"主题:{cell.font.color.theme}"
            except:
                pass
        row_dict[h] = {'value': val, 'color': color}
    all_excel_rows.append({'row_num': row_idx, 'data': row_dict})

for excel_row in all_excel_rows:
    rn = excel_row['row_num']
    d = excel_row['data']
    parts = []
    for h in headers:
        if h and h in d and d[h]['value'] is not None:
            color_mark = f"[{d[h]['color']}]" if d[h]['color'] != '默认' else ""
            parts.append(f"{h}='{str(d[h]['value'])[:30]}'{color_mark}")
    print(f"行{rn}: {' | '.join(parts)}")

# ===== 特别分析：源表20（已知是bit定义转换样例） =====
print("\n\n" + "="*140)
print("【特别分析：源表20 - bit位定义转换目标样例表】")
print("="*140)
if len(structured) > 19:
    t20 = structured[19]['table']
    paras20 = structured[19]['paras']
    print(f"前置段落: {paras20}")
    print(f"维度: {t20['row_count']}行 × {t20['col_count']}列\n")
    for row_idx, row in enumerate(t20['rows']):
        print(f"行{row_idx}: {[c['text'] for c in row]}")

# ===== 分析被跳过的表格 =====
print("\n\n" + "="*140)
print("【被跳过的表格特征分析】")
print("="*140)

skip_tables = [0, 1, 2, 12, 13, 14, 15, 16, 17, 20, 21]  # 猜测的被跳过表格
for idx in skip_tables:
    if idx < len(structured):
        item = structured[idx]
        t = item['table']
        paras = item['paras']
        all_text = ' '.join(c['text'] for row in t['rows'] for c in row)
        has_type_kw = any(kw in all_text for kw in ['UINTEGER', 'UINT', 'float', 'INT', 'bit'])
        print(f"\n源表#{idx+1}: {paras[-1][:60] if paras else '无标题'}")
        print(f"  维度: {t['row_count']}行 × {t['col_count']}列")
        print(f"  第1行: {[c['text'][:20] for c in t['rows'][0]] if t['rows'] else []}")
        print(f"  含数据类型关键字: {has_type_kw}")
        if not has_type_kw:
            print(f"  → 判断: 这是干扰表/说明表，应该被跳过")
