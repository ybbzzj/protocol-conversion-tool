# -*- coding: utf-8 -*-
"""公开问题集对应的最小回归测试。"""

import unittest
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

from backend.services.data_cleaner import DataProcessor, FormulaStandardizer, RangeValueFormatter
from backend.services.field_matcher import EnhancedFieldMatcher
from backend.services.excel_exporter import ExcelExporter
from backend.services.table_detector import (
    TableDetector,
    _extract_msg_name_from_info_row,
    parse_checksum_field_names,
)
from backend.services.table_linker import TableLinker


class QuestionRegressionTests(unittest.TestCase):
    def setUp(self):
        self.detector = TableDetector()
        self.linker = TableLinker()

    @staticmethod
    def _bit_table(table_ref, index, state_name):
        return {
            "index": index,
            "table_ref": table_ref,
            "table_type": "bit_def",
            "headers": ["位号", "状态参数"],
            "data_rows": [{"位号": "D0", "状态参数": state_name}],
        }

    def test_bit_rows_only_attach_to_exact_referenced_table(self):
        bit_tables = [
            self._bit_table("A.5", 8, "A5状态"),
            self._bit_table("A.7", 10, "A7状态"),
        ]

        unrelated = self.linker._attach_bit_rows(
            {"数据类型": "UINT16", "备注": "内容见表A.3"},
            bit_tables,
            field_table_idx=5,
            parent_headers=["内容", "数据类型", "备注"],
        )
        exact = self.linker._attach_bit_rows(
            {"数据类型": "UINT16", "备注": "内容参见表 A．5。"},
            bit_tables,
            field_table_idx=5,
            parent_headers=["内容", "数据类型", "备注"],
        )

        self.assertEqual([], unrelated)
        self.assertEqual(1, len(exact))
        self.assertEqual("A5状态", exact[0]["内容"])

    def test_placeholder_only_table_is_kept_without_fake_rows(self):
        grid = [
            ["信息名称", "某指令", "", "", "", "", ""],
            ["信源、信宿", "BCRT5-SA0-模式码0x03", "", "", "", "", ""],
            ["", "", "", "", "", "", ""],
            ["", "", "", "", "", "", ""],
            ["字序号", "内容", "类型", "格式", "值域", "单位", "数据处理方法"],
            ["—", "—", "—", "—", "—", "—", "—"],
        ]
        header_row, forced = self.detector._locate_header(grid)
        parsed = self.detector._parse_field_def_table(
            grid,
            [[False] * 7 for _ in grid],
            t_idx=4,
            preceding_para="某指令表",
            header_row_idx=header_row,
            is_config_forced=forced,
        )

        self.assertEqual(4, header_row)
        self.assertEqual("field_def", parsed["table_type"])
        self.assertEqual([], parsed["data_rows"])
        self.assertEqual("placeholder_only", parsed["empty_reason"])
        self.assertEqual("3", parsed["meta"]["数据段长度"])

    def _extract_field_rows(self, data_rows, remove_crc=True):
        headers = ["序号", "内容", "类型", "说明"]
        grid = [headers, *data_rows]
        self.detector.remove_crc_tail = remove_crc
        return self.detector._extract_data_rows(
            grid,
            headers,
            list(range(len(headers))),
            start_row=1,
            is_vmerge_cont=[[False] * len(headers) for _ in grid],
        )

    def test_checksum_before_numbered_note_is_removed(self):
        rows = [
            ["1", "有效字段", "USHORT", ""],
            ["2", "CRC校验码", "USHORT", "字号1到字号1按字节进行CRC校验"],
            ["注1：说明", "注1：说明", "注1：说明", "注1：说明"],
        ]

        removed = self._extract_field_rows(rows, remove_crc=True)
        retained = self._extract_field_rows(rows, remove_crc=False)

        self.assertEqual(["有效字段"], [row["内容"] for row in removed])
        self.assertEqual(
            ["有效字段", "CRC校验码"],
            [row["内容"] for row in retained],
        )

    def test_checksum_names_are_matched_exactly(self):
        accepted = [
            "CRC校验字", "CRC校验码", "CRC16校验字", "CRC16校验码",
            "CRC校验", "CRC16校验", "CRC校验和", "CRC16校验和", "CKS校验和",
        ]
        for checksum_name in accepted:
            with self.subTest(checksum_name=checksum_name):
                rows = self._extract_field_rows(
                    [["1", checksum_name, "USHORT", ""]],
                    remove_crc=True,
                )
                self.assertEqual([], rows)

        for ordinary_name in ("校验字", "校验码", "CRC状态", "crc校验字", "CRC 校验字"):
            with self.subTest(ordinary_name=ordinary_name):
                rows = self._extract_field_rows(
                    [["1", ordinary_name, "USHORT", ""]],
                    remove_crc=True,
                )
                self.assertEqual(ordinary_name, rows[0]["内容"])

    def test_custom_checksum_names_replace_defaults_and_match_exactly(self):
        self.detector.checksum_field_names = {"自定义校验字段"}
        rows = self._extract_field_rows(
            [
                ["1", "自定义校验字段", "USHORT", ""],
                ["2", "自定义校验字段扩展", "USHORT", ""],
                ["3", "CRC校验字", "USHORT", ""],
            ],
            remove_crc=True,
        )

        self.assertEqual(
            ["自定义校验字段扩展", "CRC校验字"],
            [row["内容"] for row in rows],
        )

    def test_checksum_field_name_request_config_is_validated(self):
        self.assertIsNone(parse_checksum_field_names(None))
        self.assertEqual([], parse_checksum_field_names("[]"))
        self.assertEqual(
            ["CRC校验字", "CKS校验和"],
            parse_checksum_field_names(
                '[" CRC校验字 ", "CRC校验字", "", "CKS校验和"]'
            ),
        )
        for invalid in ('{}', '["CRC校验字", 1]', 'not-json'):
            with self.subTest(invalid=invalid):
                with self.assertRaises(ValueError):
                    parse_checksum_field_names(invalid)

    def test_cks_reference_remark_obeys_checksum_switch(self):
        row = ["4", "CKS校验和", "USHORT", "按照规定计算，参见附录D。"]
        self.assertEqual([], self._extract_field_rows([row], remove_crc=True))
        self.assertEqual(
            "CKS校验和",
            self._extract_field_rows([row], remove_crc=False)[0]["内容"],
        )

    def test_sequence_named_field_is_not_treated_as_metadata(self):
        rows = self._extract_field_rows(
            [["1", "序号", "字", "0x0000"]],
            remove_crc=True,
        )
        self.assertEqual(1, len(rows))
        self.assertEqual("序号", rows[0]["内容"])

    def test_composite_file_headers_are_recognized_together(self):
        grid = [
            ["文件名称", "文件格式", "文件内容"],
            ["demo.bin", "BIN", "接口数据"],
        ]
        header_row, _ = self.detector._locate_header(grid)
        self.assertEqual(0, header_row)

    def test_message_name_label_is_not_returned_as_value(self):
        self.assertEqual(
            "某状态消息",
            _extract_msg_name_from_info_row(["消息名称", "某状态消息"]),
        )

    def test_single_coefficient_pair_is_standardized(self):
        standardizer = FormulaStandardizer()
        self.assertEqual(
            "0.01139x-93.33",
            standardizer.standardize("y=kx+b，其中k=0.01139，b=-93.33"),
        )

    def test_multiple_coefficient_pairs_are_joined_with_ascii_comma(self):
        text = "y=kx+b：k=0.0012205,b=-10；k=0.002441,b=-20"
        self.assertEqual(
            "0.0012205x-10,0.002441x-20",
            FormulaStandardizer().standardize(text),
        )

    def test_multiple_coefficient_pairs_are_extracted_from_data_source(self):
        processed = DataProcessor().process_row({
            "数据来源": (
                "该数据需按照y=kx+b公式进行转换，"
                "某数据系数为k=0.0012205，b=-10，"
                "另一数据系数为k=0.002441，b=-20"
            ),
        })
        self.assertEqual(
            "0.0012205x-10,0.002441x-20",
            processed["formatted"]["转换公式"],
        )

    def test_low_ambiguity_table_number_alias(self):
        match = EnhancedFieldMatcher()._alias_match("表号")
        self.assertEqual("参数表号", match["target"])

    def test_file_table_headers_map_to_selected_output_fields(self):
        matcher = EnhancedFieldMatcher(standard_fields=["名称", "内容", "数据类型"])
        self.assertEqual("名称", matcher.match_field("文件名称")["target"])
        self.assertEqual("内容", matcher.match_field("文件内容")["target"])
        self.assertEqual("数据类型", matcher.match_field("文件格式")["target"])

    def test_file_table_fields_survive_excel_export(self):
        tables = [{
            "msg_name": "某包含内容",
            "meta": {},
            "data_rows": [{
                "名称": "demo.bin",
                "内容": "接口数据",
                "数据类型": "二进制文件",
            }],
        }]
        with TemporaryDirectory() as output_dir:
            output = ExcelExporter(output_dir).export_with_template(tables, "file-table")
            ws = load_workbook(output).active
            headers = [cell.value for cell in ws[1]]
            values = [cell.value for cell in ws[2]]
            exported = dict(zip(headers, values))

        self.assertEqual("demo.bin", exported["名称"])
        self.assertEqual("接口数据", exported["内容"])
        self.assertEqual("二进制文件", exported["转换类型"])
        self.assertIsNone(exported["类型（bit）"])

    def test_parameter_code_candidates_require_manual_mapping(self):
        matcher = EnhancedFieldMatcher(standard_fields=["参数编码"])
        matcher.semantic_enabled = False
        matcher.knowledge_base.extend([
            {"source": "编号", "target": "参数编码", "confidence": 1.0},
            {"source": "信息代号", "target": "参数编码", "confidence": 1.0},
        ])

        for source in ("编号", "信息代号"):
            result = matcher.match_with_context([source])[0]
            self.assertIsNone(result["matched"])
            self.assertIn(
                "参数编码",
                [item["field"] for item in result.get("suggestions", [])],
            )

    def test_explicit_units_in_remarks_are_extracted(self):
        processor = DataProcessor()
        self.assertEqual(
            "m/s",
            processor.process_row({"备注": "单位为m/s"})["formatted"]["单位"],
        )
        self.assertEqual(
            "rad",
            processor.process_row({"备注": "单位为rad，发射系"})["formatted"]["单位"],
        )

    def test_multiple_ranges_in_value_domain_text_are_preserved(self):
        processed = DataProcessor().process_row({
            "值域": "a在地址的低16bit，a取值范围0～300、3300～3600，精度为0.1°；",
        })
        self.assertEqual(
            "[0,300],[3300,3600]",
            processed["formatted"]["值域"],
        )

    def test_hyphen_ranges_distinguish_separator_from_negative_sign(self):
        formatter = RangeValueFormatter()
        self.assertEqual("[0,100]", formatter.format_range("0-100"))
        self.assertEqual("[-40,125]", formatter.format_range("-40-125"))
        self.assertEqual("[-40,-5]", formatter.format_range("-40--5"))

    def test_free_text_value_domain_is_not_wrapped_as_formula(self):
        self.assertEqual("", RangeValueFormatter().format_range("以某某某为准"))
        processed = DataProcessor().process_row({
            "内容": "某指令",
            "类型": "UINT16",
            "值域": "具体参数及要求见附录B",
        })
        self.assertNotIn("值域", processed["formatted"])

    def test_preprocessed_type_is_stable_without_formatted_range(self):
        tables = [{
            "msg_name": "某消息",
            "meta": {},
            "data_rows": [{
                "内容": "某字段",
                "字节 类型": "USHORT",
                "类型（bit）": 16,
                "值域": "以某某某为准",
            }],
        }]
        with TemporaryDirectory() as output_dir:
            output = ExcelExporter(output_dir).export_with_template(tables, "test")
            row = list(load_workbook(output).active.iter_rows(min_row=2, max_row=2, values_only=True))[0]

        self.assertEqual(16, row[10])
        self.assertEqual("UINT16", row[11])
        self.assertIsNone(row[12])
        self.assertEqual("以某某某为准", row[15])


if __name__ == "__main__":
    unittest.main()
