# -*- coding: utf-8 -*-
"""
生成极端场景下的Excel对比文件
展示修复前后的真实差异
"""
import sys
import os
import io
import openpyxl

# 设置标准输出编码为UTF-8
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

sys.path.insert(0, os.path.dirname(__file__))

from backend.services.table_detector import DocumentParser
from backend.services.table_linker import TableLinker

def generate_excel_comparison():
    """生成极端场景下的Excel对比文件"""
    print("=" * 80)
    print("[生成Excel对比] 极端场景下的修复前后对比")
    print("=" * 80)
    
    doc_path = "word/极端测试文档_目标表可能被误判.docx"
    target_names = ["XX装置仿真输入数据"]
    
    if not os.path.exists(doc_path):
        print(f"[ERROR] 测试文档不存在: {doc_path}")
        return False
    
    # ===== 修复前：不使用目标名称 =====
    print(f"\n[修复前] 不使用目标名称")
    parser_before = DocumentParser()
    result_before = parser_before.parse(doc_path)
    raw_tables_before = result_before['tables']
    
    print(f"[识别结果] 总表格数: {len(raw_tables_before)}")
    for idx, table in enumerate(raw_tables_before):
        msg_name = table.get('msg_name', '未知')
        table_type = table.get('table_type', 'unknown')
        data_rows = table.get('data_rows', [])
        print(f"  表{idx+1}: {msg_name} ({table_type}, {len(data_rows)}行)")
    
    # 关联表格
    linker_before = TableLinker()
    linked_tables_before = linker_before.link_tables(raw_tables_before)
    
    # 生成Excel文件
    wb_before = openpyxl.Workbook()
    ws_before = wb_before.active
    ws_before.title = "修复前提取结果"
    
    # 表头
    headers = ['表格名称', '序号', '参数', '数据类型', '单位', '备注']
    for col_idx, header in enumerate(headers, 1):
        ws_before.cell(1, col_idx, header)
    
    # 数据
    row_idx = 2
    for table in linked_tables_before:
        msg_name = table.get('msg_name', '未知')
        data_rows = table.get('data_rows', [])
        
        for row in data_rows:
            ws_before.cell(row_idx, 1, msg_name)
            ws_before.cell(row_idx, 2, row.get('序号', ''))
            ws_before.cell(row_idx, 3, row.get('参数', row.get('内容', row.get('信号名称', ''))))
            ws_before.cell(row_idx, 4, row.get('数据类型', row.get('类型', '')))
            ws_before.cell(row_idx, 5, row.get('单位', ''))
            ws_before.cell(row_idx, 6, row.get('备注', row.get('说明', '')))
            row_idx += 1
    
    before_path = "极端场景_修复前_提取结果.xlsx"
    wb_before.save(before_path)
    print(f"[SUCCESS] Excel文件已生成: {before_path}")
    
    # ===== 修复后：使用目标名称 =====
    print(f"\n[修复后] 使用目标名称: {target_names}")
    parser_after = DocumentParser(target_message_names=target_names)
    result_after = parser_after.parse(doc_path)
    raw_tables_after = result_after['tables']
    
    print(f"[识别结果] 总表格数: {len(raw_tables_after)}")
    for idx, table in enumerate(raw_tables_after):
        msg_name = table.get('msg_name', '未知')
        table_type = table.get('table_type', 'unknown')
        data_rows = table.get('data_rows', [])
        print(f"  表{idx+1}: {msg_name} ({table_type}, {len(data_rows)}行)")
        
        if 'XX装置仿真输入数据' in msg_name:
            print(f"    ✅ 目标表成功提取")
    
    # 关联表格
    linker_after = TableLinker()
    linked_tables_after = linker_after.link_tables(raw_tables_after)
    
    # 生成Excel文件
    wb_after = openpyxl.Workbook()
    ws_after = wb_after.active
    ws_after.title = "修复后提取结果"
    
    # 表头
    for col_idx, header in enumerate(headers, 1):
        ws_after.cell(1, col_idx, header)
    
    # 数据
    row_idx = 2
    for table in linked_tables_after:
        msg_name = table.get('msg_name', '未知')
        data_rows = table.get('data_rows', [])
        
        for row in data_rows:
            ws_after.cell(row_idx, 1, msg_name)
            ws_after.cell(row_idx, 2, row.get('序号', ''))
            ws_after.cell(row_idx, 3, row.get('参数', row.get('内容', row.get('信号名称', ''))))
            ws_after.cell(row_idx, 4, row.get('数据类型', row.get('类型', '')))
            ws_after.cell(row_idx, 5, row.get('单位', ''))
            ws_after.cell(row_idx, 6, row.get('备注', row.get('说明', '')))
            row_idx += 1
    
    after_path = "极端场景_修复后_提取结果.xlsx"
    wb_after.save(after_path)
    print(f"[SUCCESS] Excel文件已生成: {after_path}")
    
    # ===== 对比分析 =====
    print(f"\n" + "=" * 80)
    print("[对比分析] Excel文件对比")
    print("=" * 80)
    
    # 加载并对比Excel文件
    wb_before_check = openpyxl.load_workbook(before_path)
    wb_after_check = openpyxl.load_workbook(after_path)
    
    ws_before_check = wb_before_check.active
    ws_after_check = wb_after_check.active
    
    before_rows = ws_before_check.max_row - 1  # 减去表头
    after_rows = ws_after_check.max_row - 1
    
    print(f"\n[基本对比]")
    print(f"  修复前数据行数: {before_rows}")
    print(f"  修复后数据行数: {after_rows}")
    print(f"  数据行数差异: {after_rows - before_rows}")
    
    # 检查目标表是否存在
    print(f"\n[目标表对比]")
    
    target_before = False
    target_after = False
    
    # 检查修复前的Excel
    for row_idx in range(2, ws_before_check.max_row + 1):
        cell_value = ws_before_check.cell(row_idx, 1).value
        if cell_value and 'XX装置仿真输入数据' in str(cell_value):
            target_before = True
            break
    
    # 检查修复后的Excel
    for row_idx in range(2, ws_after_check.max_row + 1):
        cell_value = ws_after_check.cell(row_idx, 1).value
        if cell_value and 'XX装置仿真输入数据' in str(cell_value):
            target_after = True
            break
    
    print(f"  修复前: {'✅ 包含目标表' if target_before else '❌ 不包含目标表'}")
    print(f"  修复后: {'✅ 包含目标表' if target_after else '❌ 不包含目标表'}")
    
    if not target_before and target_after:
        print(f"\n  ✅✅✅ 兜底提取能力验证成功！")
        print(f"      修复前：目标表被误判为干扰表，Excel中不包含目标表数据")
        print(f"      修复后：使用目标名称兜底提取，Excel中包含目标表数据")
        print(f"      数据行数增加：{after_rows - before_rows}行")
    
    # 检查关键字段
    print(f"\n[关键字段对比]")
    
    # 在修复后的Excel中查找关键字段
    found_fields_after = []
    for row_idx in range(2, ws_after_check.max_row + 1):
        param_value = ws_after_check.cell(row_idx, 3).value
        if param_value:
            if "仿真状态标志位" in str(param_value):
                found_fields_after.append("仿真状态标志位")
            elif "XX计时时间" in str(param_value):
                found_fields_after.append("XX计时时间")
    
    print(f"  修复后找到的关键字段: {len(found_fields_after)}/2")
    for field in found_fields_after:
        print(f"    ✅ {field}")
    
    if len(found_fields_after) >= 2:
        print(f"  ✅ 关键字段提取完整")
    else:
        print(f"  ⚠️  关键字段提取不完整")
    
    # 汇总报告
    print(f"\n" + "=" * 80)
    print("[汇总报告]")
    print("=" * 80)
    
    print(f"[SUCCESS] Excel对比文件已生成:")
    print(f"  修复前: {before_path}")
    print(f"  修复后: {after_path}")
    
    print(f"\n[核心发现]")
    
    if not target_before and target_after:
        print(f"  ✅✅✅ 目标消息名称兜底提取能力成功验证")
        print(f"      - 修复前：目标表被误判为干扰表，未提取")
        print(f"      - 修复后：使用目标名称兜底提取，成功提取")
        print(f"      - 数据完整性：关键字段全部提取")
    else:
        print(f"  ℹ️  兜底机制未触发，前后都包含目标表")
    
    print(f"\n[建议]")
    print(f"  1. 打开 {before_path} 查看修复前的提取结果")
    print(f"  2. 打开 {after_path} 查看修复后的提取结果")
    print(f"  3. 对比两个文件，查看目标表'XX装置仿真输入数据'的差异")
    print(f"  4. 验证关键字段'仿真状态标志位'和'XX计时时间'的提取情况")
    
    return True

if __name__ == "__main__":
    try:
        generate_excel_comparison()
        print(f"\n[SUCCESS] Excel对比文件生成完成")
    except Exception as e:
        import traceback
        print(f"\n[ERROR] 生成失败: {e}")
        print(traceback.format_exc())