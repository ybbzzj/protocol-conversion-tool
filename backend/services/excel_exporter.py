# -*- coding: utf-8 -*-
import os
import shutil
import re
import html
from datetime import datetime
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from backend.services.field_matcher import EnhancedFieldMatcher as FieldMatcher
from backend.services.data_cleaner import DataProcessor


# ── 颜色常量 ─────────────────────────────────────────────────────────────────
COLOR_RED   = "FF0000"   # 从其他字段提取的值（不确定，需人工确认）
COLOR_BLUE  = "0070C0"   # 从辅助表关联补充的元数据
COLOR_BLACK = "000000"   # 正常


def _extract_unit_from_remark(remark: str) -> Optional[str]:
    """从备注中提取单位信息"""
    if not remark:
        return None
        
    # 常见的单位提取模式（优先级从高到低）
    patterns = [
        # 最高优先级：明确标注"单位为/单位:/单位是"的形式
        r'单位[为是：:]\s*([^\s，。,\.;；]+)',
        # LSB=N单位（如 LSB=1ms）
        r'LSB\s*=\s*([\d.]+\s*(?:ms|s|μs|min|h|Hz|kHz|MHz|GHz|V|mV|A|mA|mW|W|dB|dBm|℃|°|°C|%|bit|byte|KB|MB))',
        # 单位在括号内（如"(ms)"或"[Hz]"）
        r'[（\(\[]([A-Za-z/°℃\u00b0\u2103]+)[）\)\]]',
        # 行尾的单位符号（如 "数据处理...ms"）
        r'([\wΩμ°%℃dBmVAsHzkHzMHzGHz]+)\s*$',
        # 常见单位的单词边界匹配
        r'\b(ms|μs|min|s|h|Hz|kHz|MHz|GHz|V|mV|A|mA|W|mW|dB|dBm|℃|°|°C|%|bit|byte|KB|MB)\b',
        # 组合单位（如 °/h, km/h, m/s 等）
        r'(°/[hmin]|km/h|m/s[²2\^2]?|[°℃][/]?C)',
    ]
    
    for pattern in patterns:
        match = re.search(pattern, remark, re.IGNORECASE)
        if match:
            # 获取第一个捕获组，或者整个匹配（如果没有捕获组）
            unit = match.group(1) if len(match.groups()) > 0 else match.group(0)
            # 过滤掉太长或不合理的结果
            if unit and len(unit) <= 15 and unit.strip() not in ['为', '是', '的', '', ' ']:
                return unit.strip()
    
    return None


class ExcelExporter:
    """
    Excel 导出器（重构版）

    主要改进：
    1. 主行/子行架构：主行填完整元数据+第一个字段，子行只填字段内容
    2. 值域格式：直接使用 DataProcessor 格式化后的结果（0~4294967295 格式）
    3. 颜色规则：
       - 红色：从备注/其他字段提取的值（不确定来源）
       - 蓝色：从辅助表（端口/ID）关联补充的元数据
       - 黑色：原始字段直接映射
    4. bit 子行：在对应字段的下方插入 bit 位说明子行
    """

    # 元数据键 → Excel 列名 映射
    META_TO_EXCEL: Dict[str, str] = {
        '消息ID':     'ID',
        '信息标识':   'ID',
        '信息 ID':     'ID',
        '信息名称':   '名称',
        '信源系统码': '信源系统码',
        '信源机器码': '信源机器码',
        '信宿系统码': '信宿系统码',
        '信宿机器码': '信宿机器码',
        '子地址':     '子地址',
    }

    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        self.matcher = FieldMatcher()
        self.processor = DataProcessor()

    # ── 公有接口 ──────────────────────────────────────────────────────────────

    def export_with_template(self, tables_data: List[Dict], task_id: str) -> str:
        """
        将 tables_data 写入模板 Excel 并返回输出路径。
        """
        # 使用 Config.RESOURCE_DIR 定位只读资源（如模板文件）
        from backend.config import Config
        template_path = os.path.join(Config.RESOURCE_DIR, 'word', 'csvfile', '协议模板.xlsx')
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        output_path = os.path.join(self.output_dir, f"协议_{timestamp}.xlsx")
        shutil.copy(template_path, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # 读取模板列名（第一行）
        template_headers = [
            (cell.value if cell.value else "") for cell in ws[1]
        ]
        current_row = 2

        for table in tables_data:
            msg_name = table.get('msg_name', '')
            meta     = table.get('meta', {})
            rows     = table.get('data_rows', [])

            if not rows:
                continue

            # 分离 meta 来源（用于颜色标记）
            # table.get('meta_sources', {})：键→'linked'|'original'
            meta_sources = table.get('meta_sources', {})

            first_row = True
            for row in rows:
                # ── bit 位子行 ──────────────────────────────────────────────
                if row.get('_is_bit_row'):
                    fill_data = self._build_bit_row(row, template_headers)
                    
                    # 对子行的推断类型进行类型转换处理
                    color_map = {}
                    if '_inferred_type' in row and row['_inferred_type']:
                        inferred_type = row['_inferred_type']
                        # 将推断出的类型标记为需要标红（因为是推断出来的）
                        for col_name in ['数据类型', '类型', '数据格式']:
                            if col_name in fill_data and fill_data[col_name] == inferred_type:
                                color_map[col_name] = COLOR_RED
                                break
                    
                    self._write_row(ws, current_row, template_headers, fill_data, color_map=color_map)
                    current_row += 1
                    continue

                # ── 普通字段行 ──────────────────────────────────────────────
                # 如果 row 中含有 _fmt_ 前缀的预处理结果（来自 extract.py），直接使用
                # 否则在此重新处理
                if any(k.startswith('_fmt_') for k in row):
                    # 已在 extract.py 处理过，直接拆分
                    raw_cleaned = {k: v for k, v in row.items() if not k.startswith('_fmt_') and not k.startswith('_')}
                    cleaned = raw_cleaned
                    # 还原格式化结果
                    formatted = {k[5:]: v for k, v in row.items() if k.startswith('_fmt_')}
                    # 类型信息从 row 中取（已由 extract.py 注入）
                    conv_info = {}
                    if '类型（bit）' in row:
                        conv_info['位数'] = row['类型（bit）']
                    # 从 cleaned 重新跑一次类型标准化
                    type_val = ''
                    for k, v in raw_cleaned.items():
                        if any(kw in k for kw in ['类型', 'TYPE', '数据格式', '转换类型']):
                            type_val = str(v) if v else ''
                            break
                    if type_val:
                        std_type, bits, type_status = self.processor.type_converter.convert_type(type_val)
                        conv_info['标准类型'] = std_type
                        conv_info['位数'] = bits
                        conv_info['类型状态'] = type_status  # 记录是否为推断类型
                else:
                    proc_res = self.processor.process_row(row)
                    cleaned  = proc_res['cleaned']
                    conv_info = proc_res['converted']
                    formatted = proc_res.get('formatted', {})

                fill_data = {}
                color_map = {}   # col_name → 颜色字符串

                # 1. 字段内容列（内容/参数/信号名称 → '内容'）
                content_val = self._find_content_value(cleaned)
                if content_val:
                    fill_data['内容'] = content_val

                # 2. 类型列
                # 首先检查是否有嵌套表格引用标记（"见表B.X..."）
                if row.get('_has_nested_ref'):
                    # 保留原始的嵌套引用文本，并标红
                    fill_data['转换类型'] = row.get('_nested_ref_text', '')
                    color_map['转换类型'] = COLOR_RED
                elif '标准类型' in conv_info:
                    std_type = conv_info['标准类型']
                    # 验证是否在标准类型表中
                    from backend.services.data_cleaner import DataTypeConverter
                    valid_types = set(DataTypeConverter().TYPE_MAPPING.values())
                    valid_types = {t for t, _ in valid_types} | {'ENUM', ''}  # 提取标准类型名
                    
                    if std_type in valid_types or std_type == '':
                        fill_data['转换类型'] = std_type
                        # 如果类型是推断出来的（如"字节" → UINT8），标红
                        if conv_info.get('类型状态') == 'inferred':
                            color_map['转换类型'] = COLOR_RED
                    else:
                        # 非标准类型（不在valid_types中），也填充并标红
                        fill_data['转换类型'] = std_type
                        color_map['转换类型'] = COLOR_RED
                else:
                    # 从 cleaned 数据中直接获取原始类型值，检查是否为标准类型
                    type_val = ''
                    for k, v in cleaned.items():
                        if any(kw in k for kw in ['类型', 'TYPE', '数据格式']):
                            type_val = str(v).strip() if v else ''
                            break
                    
                    if type_val:
                        # 检查是否为标准类型
                        from backend.services.data_cleaner import DataTypeConverter
                        valid_types = set(DataTypeConverter().TYPE_MAPPING.values())
                        valid_types = {t for t, _ in valid_types} | {'ENUM', ''}
                        
                        if type_val in valid_types:
                            fill_data['转换类型'] = type_val
                        else:
                            # 非标准类型，填充并标红
                            fill_data['转换类型'] = type_val
                            color_map['转换类型'] = COLOR_RED
                
                if '位数' in conv_info:
                    fill_data['类型（bit）'] = conv_info['位数']
                    # 如果类型是推断出来的，位数也标红
                    if conv_info.get('类型状态') == 'inferred':
                        color_map['类型（bit）'] = COLOR_RED

                # 3. 值域 → 判读公式列（格式化后填入）
                range_result = self._extract_range(cleaned, formatted)
                if range_result:
                    range_val, range_source = range_result
                    fill_data['判读公式'] = range_val
                    if range_source == 'extracted':
                        color_map['判读公式'] = COLOR_RED

                # 4. 单位列：先查原始单位列，为空则从备注/处理方法中提取单位符号（标红）
                unit_result = self._extract_unit(cleaned)
                if unit_result:
                    unit_val, unit_source = unit_result
                    fill_data['单位'] = unit_val
                    if unit_source == 'extracted':
                        color_map['单位'] = COLOR_RED

                # 5. 转换公式列：使用 DataProcessor 已识别并标准化的结果（aX+b 格式）
                #    DataProcessor 已按优先级从 转换公式 > 数据处理 > 数据处理方法 > 备注 中提取
                formula_val = formatted.get('转换公式', '')
                if formula_val:
                    fill_data['转换公式'] = formula_val

                # 6. 序号列
                seq_val = self._find_seq_value(cleaned)
                if seq_val:
                    fill_data['序号'] = seq_val

                # 7. 备注列：保留原始备注内容原样，不删减
                #    单位已从备注中提取到单位列（若有），备注本身不受影响
                remark_val = self._find_remark_value(cleaned)
                if remark_val:
                    fill_data['备注'] = remark_val

                # 8. 主行：注入消息名称和元数据
                if first_row:
                    # 直接使用原始消息名称，不添加序号
                    fill_data['名称'] = msg_name
                
                    # 注入元数据（来自端口表/ID 表等）
                    for mk, mv in meta.items():
                        if not mv:
                            continue
                        excel_col = self._meta_key_to_excel_col(mk, template_headers)
                        src = meta_sources.get(mk, 'original')
                        if excel_col:
                            # 不覆盖已有值（特别是名称列，msg_name 优先）
                            if excel_col in fill_data and fill_data[excel_col]:
                                continue
                            fill_data[excel_col] = mv
                            if src == 'linked':
                                color_map[excel_col] = COLOR_BLUE
                        # 不再将元数据追加到备注列，保持备注内容的原始性和完整性
                
                    first_row = False
                else:
                    # 子行：名称为空
                    fill_data['名称'] = ''

                # ── 用户手动映射（程序计算列条件覆盖，普通列直接覆盖）────
                # _override_cols 中列出的目标列由用户在前端手动映射指定。
                # 对判读公式/转换公式：只有当原始值能被标准化器正确解析时才覆盖，
                # 否则保留程序从正确源列提取并标准化的结果。
                override_cols = row.get('_override_cols') or []
                for ocol in override_cols:
                    raw_val = row.get(ocol)
                    if raw_val in (None, ''):
                        continue
                    raw_val = str(raw_val).strip()
                    if not raw_val or raw_val in ('—', '-'):
                        continue

                    if ocol == '判读公式':
                        # 只有当值能被解析为有效范围/枚举时才覆盖
                        from backend.services.data_cleaner import RangeValueFormatter
                        formatted_range = RangeValueFormatter().format_range(raw_val)
                        # 有效范围格式: [数字,数字] 或 {枚举值}
                        if (re.match(r'^\[[\d\s,xa-fA-F\-]+\]$', formatted_range)
                                or re.match(r'^\{[^}]+\}$', formatted_range)):
                            fill_data['判读公式'] = formatted_range
                        # 否则保留程序从值域列提取的结果

                    elif ocol == '转换公式':
                        # 只有当值能被解析为有效公式时才覆盖
                        from backend.services.data_cleaner import FormulaStandardizer
                        formatted_formula = FormulaStandardizer().standardize(raw_val)
                        # 有效公式格式: 包含 X（aX+b 形式）
                        if re.search(r'[\d.]+[xX]', formatted_formula):
                            fill_data['转换公式'] = formatted_formula
                        # 否则保留程序从数据处理方法等列提取的结果

                    elif ocol in template_headers:
                        fill_data[ocol] = raw_val

                self._write_row(ws, current_row, template_headers, fill_data, color_map)
                current_row += 1

        wb.save(output_path)
        return output_path

    # ── 行写入 ────────────────────────────────────────────────────────────────

    def _write_row(self, ws, row_num: int, template_headers: List[str],
                   fill_data: Dict, color_map: Dict):
        """将 fill_data 写入 ws 的 row_num 行，按 template_headers 对齐列"""
        for col_idx, col_name in enumerate(template_headers, 1):
            if not col_name:
                continue
            val = self._find_value_for_column(col_name, fill_data)
            if val is None:
                continue
            cell = ws.cell(row=row_num, column=col_idx, value=val)

            # 颜色标记
            color = color_map.get(col_name)
            if color:
                cell.font = Font(color=color)

    # ── 辅助方法 ──────────────────────────────────────────────────────────────

    def _build_bit_row(self, bit_row: Dict, template_headers: List[str]) -> Dict:
        """
        将 bit 子行转换为 fill_data 字典。
        
        新的子行结构保留了父表的表头，需要正确映射字段值。
        子行中的推断类型（_inferred_type）会被填充到数据类型列。
        """
        fill_data = {}
        
        # 主行的"名称"列为空（子行没有独立的名称）
        fill_data['名称'] = ''
        
        # 遍历所有字段，将子行的值复制到fill_data中
        for key, value in bit_row.items():
            # 跳过内部字段（除了_inferred_type）
            if key.startswith('_') and key != '_inferred_type':
                continue
            
            # 处理推断出的类型
            if key == '_inferred_type' and value:
                # 将推断出的类型填充到数据类型列
                # 查找合适的类型列
                for col_name in ['数据类型', '类型', '数据格式']:
                    if col_name in bit_row and bit_row[col_name] == '':
                        fill_data[col_name] = value
                        break
                continue
            
            # 直接复制字段值
            if value is not None and value != '':
                fill_data[key] = value
        
        # 兼容旧的"子内容"字段（如果新字段未生成）
        if '子内容' in bit_row and '内容' not in fill_data and '数据含义' not in fill_data:
            fill_data['内容'] = bit_row.get('子内容', '')
        
        return fill_data

    def _find_content_value(self, cleaned: Dict) -> str:
        """从清洗数据中提取内容/参数/信号名称列的值"""
        candidates = ['内容', '数据含义', '字段', '参数', '信号名称', '名称']
        for k in cleaned:
            for cand in candidates:
                if cand in k:
                    v = cleaned[k]
                    if v and v not in ('—', '-'):
                        return str(v).strip()
        return ''

    def _find_seq_value(self, cleaned: Dict) -> str:
        """提取序号"""
        for k, v in cleaned.items():
            if '序号' in k or k == 'NO':
                if v and v not in ('—', '-'):
                    return str(v).strip()
        return ''

    def _find_remark_value(self, cleaned: Dict) -> str:
        """
        提取备注/说明内容，保留原始内容原样。
        
        查找顺序（只取第一个有值的字段）：
        1. 备注 / 说明 / 数据来源 列
        2. 数据处理方法 列（若无专用备注列，把处理方法说明放入备注）
        """
        # 优先查专用备注字段
        for k, v in cleaned.items():
            if any(kw in k for kw in ['备注', '说明', '数据来源']):
                if v and str(v).strip() not in ('—', '-', ''):
                    return str(v).strip()
        # 其次查数据处理方法（作为补充说明）
        for k, v in cleaned.items():
            if '数据处理方法' in k:
                if v and str(v).strip() not in ('—', '-', ''):
                    return str(v).strip()
        return ''

    def _extract_range(self, cleaned: Dict, formatted: Dict) -> Optional[tuple]:
        """
        提取值域，返回 (格式化后的字符串, 来源)。
        来源：'original'（直接来自值域列）、'extracted'（从备注等提取，标红）
        """
        # 优先使用 DataProcessor 已格式化的值域
        if '值域' in formatted:
            return formatted['值域'], 'original'

        # 原始值域/取值范围/区间列
        for k, v in cleaned.items():
            if any(kw in k for kw in ['值域', '取值范围', '区间']):
                val = str(v).strip() if v else ''
                if val and val not in ('—', '-', ''):
                    from backend.services.data_cleaner import RangeValueFormatter
                    return RangeValueFormatter().format_range(val), 'original'

        # 从备注中提取
        remark = ''
        for k, v in cleaned.items():
            if any(kw in k for kw in ['备注', '说明']):
                remark = str(v) if v else ''
                break

        if remark:
            # 按优先级从高到低提取范围，防止误识别（如 "LSB=1ms" 不是范围）
            range_patterns = [
                # 最高优先级：明确标注"取值范围"或"值域"
                r'取值范围[：:]*\s*([\dxXa-fA-F]+\s*[~\-]\s*[\dxXa-fA-F]+)',
                r'值域[：:]*\s*\[?([\dxXa-fA-F]+\s*[~\-]\s*[\dxXa-fA-F]+)\]?',
                # 16进制范围（如 "0~0xFFFF"）
                r'(\d+\s*[~\-]\s*0x[0-9A-Fa-f]+)',
                r'(0x[0-9A-Fa-f]+\s*[~\-]\s*0x[0-9A-Fa-f]+)',
                # 纯数字范围（如 "0~400"），要求是整行或句子开头，或明确上下文
                # 避免匹配 "1ms" 中的 "1" 这类非范围内容
                r'^(0\s*[~\-]\s*\d+)$',        # 整行就是范围
                r'^(\d+\s*[~\-]\s*\d+)$',       # 整行就是范围
            ]
            for pattern in range_patterns:
                m = re.search(pattern, remark.strip(), re.MULTILINE)
                if m:
                    extracted = m.group(1).strip()
                    # 过滤太短的匹配（单个数字不是范围）
                    if extracted and re.search(r'[~\-]', extracted):
                        from backend.services.data_cleaner import RangeValueFormatter
                        return RangeValueFormatter().format_range(extracted), 'extracted'
            
            # 提取枚举值格式（如 "0x1701:供电 0x1702:断电"）
            enum_patterns = [
                # 枚举值格式：0x1701:供电 0x1702:断电
                # 支持多种冒号：英文冒号 : (U+003A)、中文冒号：(U+FF1A)、全角冒号：(U+FE55)、特殊冒号︓(U+FE3A)
                # 不要求空格，多个枚举项之间可以有或没有分隔符
                r'0x[0-9A-Fa-f]+\s*[:：︓︓]\s*[^\s,，;；.。]+',
                # 枚举值格式：{0x1701, 0x1702}
                r'\{([^\}]+)\}',
            ]
            for pattern in enum_patterns:
                if pattern.startswith(r'0x'):
                    # 对于 0xNNN:描述 格式，使用 findall 获取所有匹配项
                    matches = re.findall(pattern, remark.strip())
                    if len(matches) >= 2:
                        # 提取所有十六进制值（保持原样）
                        hex_values = []
                        for match in matches:
                            hex_match = re.search(r'0x([0-9A-Fa-f]+)', match, re.IGNORECASE)
                            if hex_match:
                                hex_values.append('0x' + hex_match.group(1))
                        
                        if len(hex_values) >= 2:
                            enum_str = '{' + ', '.join(hex_values) + '}'
                            return enum_str, 'extracted'
                else:
                    # 对于 {} 格式，使用 search
                    m = re.search(pattern, remark.strip())
                    if m:
                        extracted = m.group(0).strip()
                        # 处理枚举值
                        if extracted:
                            # 转换为 {} 格式
                            enum_values = []
                            # 处理逗号分隔格式
                            if ',' in extracted:
                                items = extracted.replace('{', '').replace('}', '').split(',')
                                for item in items:
                                    val = item.strip()
                                    # 保持原始进制
                                    enum_values.append(val)
                            # 如果有枚举值，返回 {} 格式
                            if enum_values:
                                enum_str = '{' + ', '.join(enum_values) + '}'
                                return enum_str, 'extracted'

        return None

    def _extract_unit(self, cleaned: Dict) -> Optional[tuple]:
        """
        提取单位，返回 (单位字符串, 来源)。
        来源：'original'（来自单位列）、'extracted'（从备注中提取，标红）

        规则：
        - 优先从单位列提取（黑色，正常）
        - 单位列为空时，从备注/说明/数据处理列中识别单位符号（红色，标注来源不确定）
        - 备注列内容始终保留原样，不会因为单位提取而被删改
        """
        # 优先：单位列直接提取
        for k, v in cleaned.items():
            if '单位' in k and '长度' not in k:
                if v and str(v).strip() not in ('—', '-', '无', ''):
                    return str(v).strip(), 'original'

        # 备选：从备注/说明/数据处理方法列中识别单位符号（备注保留原样，只读取）
        search_text = ''
        for k, v in cleaned.items():
            if any(kw in k for kw in ['备注', '说明', '数据处理']):
                if v and str(v).strip() not in ('—', '-', ''):
                    search_text += str(v) + ' '

        if not search_text.strip():
            return None

        # 按优先级从高到低匹配，越具体的单位越靠前
        unit_patterns = [
            # 组合单位（最高优先级，在单个字符单位前提取）
            (r'(?<![a-zA-Z\d])km/h(?![a-zA-Z])',  'km/h'),
            (r'(?<![a-zA-Z\d])m/s[²2\^2]',        'm/s²'),
            (r'(?<![a-zA-Z\d])m/s(?![²2a-zA-Z])', 'm/s'),
            (r'[°∠]\s*/\s*h\b',                    '°/h'),    # 角速率
            (r'[°∠]\s*/\s*min\b',                  '°/min'),  # 角速率/分钟
            (r'[°∠]\s*/\s*s\b',                    '°/s'),    # 角速率/秒
            # LSB 相关单位（从"LSB=1ms"中提取，需要捕获 °/h 等组合单位）
            (r'LSB\s*=\s*[\d.]*\s*(?:ms|μs|s|°/h|km/h|m/s)',  'ms'),  # 这里简化处理
            # 频率单位
            (r'\b(kHz|千赫)\b',            'kHz'),
            (r'\b(MHz|兆赫)\b',            'MHz'),
            (r'\b(GHz)\b',                'GHz'),
            (r'\b(Hz|赫兹)\b',             'Hz'),
            # 毫秒，需要特别小心与其他单位混淆
            (r'(?<![a-zA-Z])ms(?![a-zA-Z\d])', 'ms'),
            # 微秒
            (r'(?<![a-zA-Z])μs\b',        'μs'),
            (r'(?<![a-zA-Z])us\b',        'μs'),     # us 也表示微秒
            # 温度
            (r'(℃|°C|摄氏度)',             '℃'),
            # 角度（仅在不是斜杠形式的时候提取）
            (r'(?<![/分])°(?![/hmsC])',    '°'),
            # 电压
            (r'\b(mV|毫伏)\b',             'mV'),
            (r'(?<![a-zA-Z])V(?![a-zA-Z])', 'V'),
            # 电流
            (r'\b(mA|毫安)\b',             'mA'),
            (r'(?<![a-zA-Z])A(?![a-zA-Z])', 'A'),
            # 功率
            (r'\b(mW|毫瓦)\b',             'mW'),
            (r'\b(W|瓦)\b',                'W'),
            # 增益
            (r'\b(dB|分贝)\b',             'dB'),
            (r'\b(dBm)\b',                 'dBm'),
            # 数据大小
            (r'\b(bit|位)\b',              'bit'),
            (r'\b(byte|字节)\b',           'byte'),
            (r'\b(KB)\b',                  'KB'),
            (r'\b(MB)\b',                  'MB'),
            # 时间单位
            (r'(?<![a-zA-Z])min\b',        'min'),   # 分钟，要求 min 后边界清晰
            (r'\b(h|小时)\b',              'h'),
            (r'(?<![a-zA-Z])s(?![a-zA-Z])', 's'),   # 秒，排除其他字母组合
            # 百分比
            (r'(%)',                        '%'),
            # 欧姆（电阻）
            (r'(Ω|ohm)',                   'Ω'),
        ]
        for pattern, unit_name in unit_patterns:
            if re.search(pattern, search_text, re.IGNORECASE):
                return unit_name, 'extracted'

        return None

    def _meta_key_to_excel_col(self, meta_key: str, template_headers: List[str]) -> Optional[str]:
        """
        将元数据键名映射到 Excel 列名。
        先查精确映射表，再做模糊匹配。
        模糊匹配要求列名是元数据键的子串（而非反过来），
        避免短列名（如'名称'）被长元数据键（如'上级信息名称'）误匹配。
        """
        # 精确映射
        col = self.META_TO_EXCEL.get(meta_key)
        if col and col in template_headers:
            return col

        # 模糊匹配：列名等于元数据键，或列名是元数据键的子串
        # 但要求列名长度 >= 3，避免短列名（如'名称'）匹配到包含它的长元数据键（如'上级信息名称'）
        for h in template_headers:
            if h and h == meta_key:
                return h
            if h and len(h) >= 3 and h in meta_key:
                return h

        return None

    def _find_value_for_column(self, col_name: str, fill_data: Dict) -> Any:
        """根据 Excel 列名从 fill_data 中查找对应值（精确+别名匹配）"""
        if col_name in fill_data:
            return fill_data[col_name]

        # 通过 FieldMatcher 查找别名
        for k, v in fill_data.items():
            res = self.matcher.match_field(k)
            target = res.get('target') if isinstance(res, dict) else getattr(res, 'target', None)
            if target == col_name:
                return v

        # 手动补丁
        if col_name == '内容':
            return fill_data.get('参数', fill_data.get('信号名称', fill_data.get('数据含义', None)))
        if col_name == '转换类型':
            return fill_data.get('数据类型', fill_data.get('类型', None))
        if col_name == '判读公式':
            return fill_data.get('判读公式', None)

        return None

    # ── 保留旧接口（兼容外部调用）────────────────────────────────────────────

    def _clean_html_entities_in_string(self, text: str) -> str:
        if not text or not isinstance(text, str):
            return text
        text = html.unescape(text)
        text = re.sub(r'<[^>]+>', '', text)
        text = ''.join(
            c for c in text
            if (32 <= ord(c) <= 126)
            or ('\u4e00' <= c <= '\u9fff')
            or c in '→→→è-_.:：（）【】、，。'
        )
        return text.strip()

    def _parse_source_destination(self, combined_value: str, available_columns: List[str]) -> dict:
        """兼容旧版解析信源信宿复合字段"""
        result = {}
        if not combined_value:
            return result
        for sep in ['→', 'è', '-']:
            if sep in combined_value:
                parts = combined_value.split(sep, 1)
                if len(parts) == 2:
                    src = self._clean_html_entities_in_string(parts[0]).strip()
                    dst = self._clean_html_entities_in_string(parts[1]).strip()
                    if src and dst:
                        if '信源机器码' in available_columns:
                            result['信源机器码'] = src
                        if '信宿机器码' in available_columns:
                            result['信宿机器码'] = dst
                        return result
        return result
