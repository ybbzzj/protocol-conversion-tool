#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
验证修复效果：检查 Excel 中的备注是否还有多余内容
"""
import sys
import os
import glob

sys.path.insert(0, 'backend')

from openpyxl import load_workbook

files = sorted(glob.glob('tests/backend/outputs/*.xlsx'))
if not files:
    print("未找到 Excel 文件")
    sys.exit(1)

latest_file = files[-1]
print(f"检查最新文件：{os.path.basename(latest_file)}\n")

wb = load_workbook(latest_file)
ws = wb.active

headers = [cell.value for cell in ws[1]]
print(f"表头：{headers}\n")

suspicious_keywords = ['发起时机', '错误处理', '传输周期']
found_issues = False

for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
    for col_idx, value in enumerate(row):
        if value is not None and str(value).strip():
            header_name = headers[col_idx] if col_idx < len(headers) else f"列{col_idx+1}"
            
            # 检查备注列
            if header_name == '备注':
                value_str = str(value)
                for keyword in suspicious_keywords:
                    if keyword in value_str:
                        print(f"⚠ 行 {row_idx} 备注意图发现可疑内容:")
                        print(f"   {header_name}: {value}")
                        print(f"   包含关键词：{keyword}\n")
                        found_issues = True

if not found_issues:
    print("✓ 验证通过！所有备注字段都没有多余的元数据内容。")
else:
    print("✗ 验证失败！仍有备注包含多余内容。")
