# -*- coding: utf-8 -*-
"""
生成修复前后对比的Excel文件
使用真实的代码生成实际的xlsx文件
"""
import sys
import os
import io
# 设置标准输出编码为UTF-8
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

sys.path.insert(0, os.path.dirname(__file__))

from backend.services.table_detector import DocumentParser
from backend.services.table_linker import TableLinker
from backend.services.excel_exporter import ExcelExporter

def generate_before_fix_excel():
    """生成修复前的Excel文件（不使用目标名称）"""
    print("=" * 80)
    print("[修复前] 生成Excel文件（不使用目标名称）")
    print("=" * 80)
    
    test_doc = "word/混合模式协议(公开).docx"
    
    if not os.path.exists(test_doc):
        print(f"[ERROR] 测试文档不存在: {test_doc}")
        return None
    
    # 使用修复前的逻辑：不传入目标名称
    parser = DocumentParser()  # 不传入target_message_names
    result = parser.parse(test_doc)
    raw_tables = result['tables']
    
    print(f"[识别结果] 总表格数: {len(raw_tables)}")
    for idx, table in enumerate(raw_tables):
        print(f"  表{idx+1}: {table.get('msg_name', '未知')} ({table.get('table_type', 'unknown')})")
    
    # 关联表格
    linker = TableLinker()
    linked_tables = linker.link_tables(raw_tables)
    
    print(f"[关联结果] 关联表格数: {len(linked_tables)}")
    
    # 导出Excel - 使用固定文件名
    output_dir = os.path.dirname(__file__)
    output_path = os.path.join(output_dir, "修复前_提取结果.xlsx")
    
    # 手动创建Excel文件（不使用模板）
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "提取结果"
        
        # 写入表头
        headers = ['表格名称', '序号', '参数', '数据类型', '单位', '备注']
        for col_idx, header in enumerate(headers, 1):
            ws.cell(1, col_idx, header)
        
        # 写入数据
        row_idx = 2
        for table in linked_tables:
            msg_name = table.get('msg_name', '未知')
            data_rows = table.get('data_rows', [])
            
            for row in data_rows:
                ws.cell(row_idx, 1, msg_name)
                ws.cell(row_idx, 2, row.get('序号', ''))
                ws.cell(row_idx, 3, row.get('参数', row.get('内容', row.get('信号名称', ''))))
                ws.cell(row_idx, 4, row.get('数据类型', row.get('类型', '')))
                ws.cell(row_idx, 5, row.get('单位', ''))
                ws.cell(row_idx, 6, row.get('备注', row.get('说明', '')))
                row_idx += 1
        
        wb.save(output_path)
        print(f"[SUCCESS] Excel文件已生成: {output_path}")
        
        # 显示文件信息
        if os.path.exists(output_path):
            file_size = os.path.getsize(output_path)
            print(f"[文件信息] 文件大小: {file_size} bytes")
            print(f"[文件信息] 数据行数: {row_idx - 2}行")
        
        return output_path
    except Exception as e:
        print(f"[ERROR] Excel导出失败: {e}")
        import traceback
        print(traceback.format_exc())
        return None

def generate_after_fix_excel():
    """生成修复后的Excel文件（使用目标名称）"""
    print("\n" + "=" * 80)
    print("[修复后] 生成Excel文件（使用目标名称）")
    print("=" * 80)
    
    test_doc = "word/混合模式协议(公开).docx"
    target_names = ["XX装置仿真输入数据"]
    
    if not os.path.exists(test_doc):
        print(f"[ERROR] 测试文档不存在: {test_doc}")
        return None
    
    # 使用修复后的逻辑：传入目标名称
    parser = DocumentParser(target_message_names=target_names)
    result = parser.parse(test_doc)
    raw_tables = result['tables']
    
    print(f"[识别结果] 总表格数: {len(raw_tables)}")
    for idx, table in enumerate(raw_tables):
        print(f"  表{idx+1}: {table.get('msg_name', '未知')} ({table.get('table_type', 'unknown')})")
    
    # 关联表格
    linker = TableLinker()
    linked_tables = linker.link_tables(raw_tables)
    
    print(f"[关联结果] 关联表格数: {len(linked_tables)}")
    
    # 导出Excel - 使用固定文件名
    output_dir = os.path.dirname(__file__)
    output_path = os.path.join(output_dir, "修复后_提取结果.xlsx")
    
    # 手动创建Excel文件（不使用模板）
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "提取结果"
        
        # 写入表头
        headers = ['表格名称', '序号', '参数', '数据类型', '单位', '备注']
        for col_idx, header in enumerate(headers, 1):
            ws.cell(1, col_idx, header)
        
        # 写入数据
        row_idx = 2
        for table in linked_tables:
            msg_name = table.get('msg_name', '未知')
            data_rows = table.get('data_rows', [])
            
            for row in data_rows:
                ws.cell(row_idx, 1, msg_name)
                ws.cell(row_idx, 2, row.get('序号', ''))
                ws.cell(row_idx, 3, row.get('参数', row.get('内容', row.get('信号名称', ''))))
                ws.cell(row_idx, 4, row.get('数据类型', row.get('类型', '')))
                ws.cell(row_idx, 5, row.get('单位', ''))
                ws.cell(row_idx, 6, row.get('备注', row.get('说明', '')))
                row_idx += 1
        
        wb.save(output_path)
        print(f"[SUCCESS] Excel文件已生成: {output_path}")
        
        # 显示文件信息
        if os.path.exists(output_path):
            file_size = os.path.getsize(output_path)
            print(f"[文件信息] 文件大小: {file_size} bytes")
            print(f"[文件信息] 数据行数: {row_idx - 2}行")
        
        return output_path
    except Exception as e:
        print(f"[ERROR] Excel导出失败: {e}")
        import traceback
        print(traceback.format_exc())
        return None

def compare_excel_files(before_path, after_path):
    """对比两个Excel文件"""
    print("\n" + "=" * 80)
    print("[对比分析] 修复前后Excel文件对比")
    print("=" * 80)
    
    if not before_path or not after_path:
        print("[ERROR] Excel文件生成失败，无法对比")
        return
    
    try:
        import openpyxl
        
        # 加载修复前的Excel
        wb_before = openpyxl.load_workbook(before_path)
        sheets_before = wb_before.sheetnames
        
        print(f"\n[修复前Excel]")
        print(f"  工作表数量: {len(sheets_before)}")
        print(f"  工作表列表: {', '.join(sheets_before)}")
        
        for sheet_name in sheets_before:
            ws = wb_before[sheet_name]
            rows = ws.max_row
            cols = ws.max_column
            print(f"  {sheet_name}: {rows}行 x {cols}列")
            
            # 显示前几行数据
            if rows > 0:
                print(f"    前3行数据:")
                for row_idx in range(1, min(4, rows+1)):
                    row_data = []
                    for col_idx in range(1, min(6, cols+1)):
                        cell_value = ws.cell(row_idx, col_idx).value
                        if cell_value:
                            row_data.append(str(cell_value)[:20])
                    if row_data:
                        print(f"      行{row_idx}: {', '.join(row_data)}")
        
        # 加载修复后的Excel
        wb_after = openpyxl.load_workbook(after_path)
        sheets_after = wb_after.sheetnames
        
        print(f"\n[修复后Excel]")
        print(f"  工作表数量: {len(sheets_after)}")
        print(f"  工作表列表: {', '.join(sheets_after)}")
        
        for sheet_name in sheets_after:
            ws = wb_after[sheet_name]
            rows = ws.max_row
            cols = ws.max_column
            print(f"  {sheet_name}: {rows}行 x {cols}列")
            
            # 显示前几行数据
            if rows > 0:
                print(f"    前3行数据:")
                for row_idx in range(1, min(4, rows+1)):
                    row_data = []
                    for col_idx in range(1, min(6, cols+1)):
                        cell_value = ws.cell(row_idx, col_idx).value
                        if cell_value:
                            row_data.append(str(cell_value)[:20])
                    if row_data:
                        print(f"      行{row_idx}: {', '.join(row_data)}")
        
        # 对比分析
        print(f"\n[对比结果]")
        
        # 工作表数量对比
        if len(sheets_after) > len(sheets_before):
            print(f"  ✅ 工作表数量增加: {len(sheets_before)} → {len(sheets_after)}")
        elif len(sheets_after) == len(sheets_before):
            print(f"  ℹ️  工作表数量相同: {len(sheets_before)}")
        else:
            print(f"  ⚠️  工作表数量减少: {len(sheets_before)} → {len(sheets_after)}")
        
        # 检查目标表是否存在
        target_sheet = "XX装置仿真输入数据"
        target_before = target_sheet in sheets_before
        target_after = target_sheet in sheets_after
        
        if target_before and target_after:
            print(f"  ✅ 目标表前后都存在: {target_sheet}")
            
            # 对比目标表的数据行数
            ws_before_target = wb_before[target_sheet]
            ws_after_target = wb_after[target_sheet]
            
            rows_before = ws_before_target.max_row
            rows_after = ws_after_target.max_row
            
            print(f"  📊 目标表数据行数: {rows_before} → {rows_after}")
            
            # 检查关键字段
            print(f"  🔍 检查关键字段:")
            
            # 在修复后的Excel中查找关键字段
            found_fields = []
            for row_idx in range(1, rows_after+1):
                for col_idx in range(1, ws_after_target.max_column+1):
                    cell_value = ws_after_target.cell(row_idx, col_idx).value
                    if cell_value:
                        if "仿真状态标志位" in str(cell_value):
                            found_fields.append("仿真状态标志位")
                            # 查找同一行的数据类型
                            for check_col in range(1, ws_after_target.max_column+1):
                                type_value = ws_after_target.cell(row_idx, check_col).value
                                if type_value and "UCHAR" in str(type_value):
                                    print(f"    ✅ 找到: 仿真状态标志位 / UCHAR")
                                    break
                        elif "XX计时时间" in str(cell_value):
                            found_fields.append("XX计时时间")
                            # 查找同一行的数据类型
                            for check_col in range(1, ws_after_target.max_column+1):
                                type_value = ws_after_target.cell(row_idx, check_col).value
                                if type_value and "UINTEGER-32" in str(type_value):
                                    print(f"    ✅ 找到: XX计时时间 / UINTEGER-32")
                                    break
            
            if len(found_fields) >= 2:
                print(f"  ✅ 关键字段提取完整: {len(found_fields)}/2")
            else:
                print(f"  ⚠️  关键字段提取不完整: {len(found_fields)}/2")
        
        elif target_after:
            print(f"  ✅ 修复后找到目标表: {target_sheet}")
            print(f"  ❌ 修复前未找到目标表")
        else:
            print(f"  ❌ 前后都未找到目标表: {target_sheet}")
        
        print(f"\n[SUCCESS] Excel对比分析完成")
        
    except ImportError:
        print("[WARNING] openpyxl未安装，无法详细对比Excel内容")
        print("[INFO] 请手动打开Excel文件查看对比")
    except Exception as e:
        print(f"[ERROR] Excel对比失败: {e}")
        import traceback
        print(traceback.format_exc())

def main():
    """主函数"""
    print("=" * 80)
    print("[生成Excel对比文件] 修复前后真实Excel文件对比")
    print("=" * 80)
    
    # 生成修复前的Excel
    before_path = generate_before_fix_excel()
    
    # 生成修复后的Excel
    after_path = generate_after_fix_excel()
    
    # 对比两个Excel文件
    compare_excel_files(before_path, after_path)
    
    # 生成汇总报告
    print("\n" + "=" * 80)
    print("[汇总报告]")
    print("=" * 80)
    
    if before_path and after_path:
        print(f"[SUCCESS] 成功生成对比Excel文件:")
        print(f"  修复前: {before_path}")
        print(f"  修复后: {after_path}")
        print(f"\n[建议] 请打开Excel文件查看详细对比:")
        print(f"  1. 打开 {before_path} 查看修复前的提取结果")
        print(f"  2. 打开 {after_path} 查看修复后的提取结果")
        print(f"  3. 对比两个文件的工作表数量、数据行数、字段完整性")
        print(f"  4. 特别关注目标表 'XX装置仿真输入数据' 的提取情况")
    else:
        print(f"[ERROR] Excel文件生成失败")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        import traceback
        print(f"\n[ERROR] 执行失败: {e}")
        print(traceback.format_exc())