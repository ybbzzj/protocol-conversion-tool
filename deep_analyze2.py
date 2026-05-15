#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
深度分析Part2：
1. 完整打印每个源表的合并单元格信息
2. 分析每个源表的"角色"（端口表/字段定义表/状态表/bit表等）
3. 精准分析Excel输出的颜色信息
4. 分析哪些表被提取/哪些被忽略
"""
from docx import Document
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from openpyxl import load_workbook

DOCX_PATH = '/Users/yuanyuqing/Documents/code/schoolProject/word/测试协议20260227.docx'
XLSX_PATH = '/Users/yuanyuqing/Documents/code/schoolProject/word/csvfile/转换结果20260227.xlsx'

doc = Document(DOCX_PATH)

# ===== 重建文档结构（段落+表格，保留顺序和前置标题）=====
elements = []
for el in doc.element.body:
    if isinstance(el, CT_P):
        ns = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'
        t_nodes = el.findall(f'.//{{{ns}}}t')
        text = ''.join((t.text or '') for t in t_nodes).strip()
        if text:
            elements.append({'type': 'para', 'text': text})
    elif isinstance(el, CT_Tbl):
        table = Table(el, doc)
        ns = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'
        rows_data = []
        for row in table.rows:
            row_data = []
            for cell in row.cells:
                tc = cell._tc
                text = cell.text.strip()
                gs_node = tc.find(f'.//{{{ns}}}gridSpan')
                gridSpan = int(gs_node.get(f'{{{ns}}}val')) if gs_node is not None and gs_node.get(f'{{{ns}}}val') else 1
                vm_node = tc.find(f'.//{{{ns}}}vMerge')
                if vm_node is not None:
                    v = vm_node.get(f'{{{ns}}}val')
                    vMerge = v if v else 'continue'
                else:
                    vMerge = None
                row_data.append({'text': text, 'gridSpan': gridSpan, 'vMerge': vMerge})
            rows_data.append(row_data)
        elements.append({
            'type': 'table',
            'rows': rows_data,
            'row_count': len(table.rows),
            'col_count': len(table.columns)
        })

# 重建：为每个表格附加前置标题
structured = []
last_paras = []
for elem in elements:
    if elem['type'] == 'para':
        last_paras.append(elem['text'])
    elif elem['type'] == 'table':
        structured.append({
            'context_paras': list(last_paras),
            'table': elem
        })
        last_paras = []

print("="*140)
print("【完整源表详细分析（含合并单元格、上下文标题）】")
print("="*140)

for table_idx, item in enumerate(structured):
    t = item['table']
    paras = item['context_paras']
    print(f"\n{'='*140}")
    print(f"【源表 #{table_idx+1}】")
    print(f"  前置段落: {paras}")
    print(f"  维度: {t['row_count']} 行 × {t['col_count']} 列")
    print(f"{'='*140}")
    for row_idx, row in enumerate(t['rows']):
        cells_display = []
        for c in row:
            info = f"'{c['text'][:30]}'"
            if c['gridSpan'] > 1:
                info += f"[横跨{c['gridSpan']}列]"
            if c['vMerge']:
                info += f"[竖合并:{c['vMerge']}]"
            cells_display.append(info)
        print(f"  行{row_idx:2d}: {cells_display}")

print(f"\n\n总源表数: {len(structured)}")

# ===== Excel精确颜色分析 =====
print("\n\n" + "="*140)
print("【Excel颜色精确分析】")
print("="*140)

wb = load_workbook(XLSX_PATH)
ws = wb.active
headers = [cell.value for cell in ws[1]]

print(f"\n列名: {headers}\n")

for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
    row_parts = []
    for col_idx, cell in enumerate(row):
        if cell.value is None:
            continue
        # 精确获取颜色
        color_str = "黑色(默认)"
        if cell.font and cell.font.color:
            try:
                t = cell.font.color.type
                if t == 'rgb':
                    rgb = cell.font.color.rgb
                    if rgb and rgb not in ('00000000', 'FF000000'):
                        color_str = f"RGB:{rgb}"
                elif t == 'theme':
                    color_str = f"主题色:{cell.font.color.theme}"
                elif t == 'indexed':
                    color_str = f"索引色:{cell.font.color.indexed}"
            except Exception as e:
                color_str = f"解析失败:{e}"
        row_parts.append(f"  {headers[col_idx]}='{str(cell.value)[:40]}'  [{color_str}]")
    if row_parts:
        print(f"\n行{row_idx}:")
        for p in row_parts:
            print(p)

# ===== 分析哪些源表被忽略 =====
print("\n\n" + "="*140)
print("【源表角色分析：哪些被提取，哪些被忽略】")
print("="*140)

# 找出Excel里出现的名称
excel_names = set()
excel_contents = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        excel_names.add(str(row[0]))
    if row[8]:
        excel_contents.add(str(row[8]))

print(f"\nExcel中出现的【名称】: {sorted(excel_names)}")
print(f"\nExcel中出现的【内容】: {sorted(excel_contents)}")

print("\n\n【各源表角色判断】")
for table_idx, item in enumerate(structured):
    t = item['table']
    paras = item['context_paras']
    all_text = ' '.join(c['text'] for row in t['rows'] for c in row)
    
    # 判断表类型
    role = "❓ 未知"
    if any('消息ID' in c['text'] or '0x8' in c['text'] for row in t['rows'] for c in row):
        if t['col_count'] <= 5:
            role = "📋 端口映射表 (定义消息ID)"
    if any('数据含义' in c['text'] or '数据类型' in c['text'] for row in t['rows'] for c in row):
        # 看是否有关键内容在Excel中
        has_match = any(c['text'] in excel_contents for row in t['rows'] for c in row)
        role = "📊 字段定义表 (被提取)" if has_match else "🚫 字段定义表 (可能被忽略)"
    if any('bit' in c['text'].lower() or '状态' in c['text'] for row in t['rows'] for c in row) and t['col_count'] <= 4:
        role = "🔧 bit位定义表"
    if all(any(kw in c['text'] for c in row) or all(c['text'] == '' for c in row) for row in t['rows'] for _ in [None]) if False else False:
        role = "💬 说明文字/干扰表"
    
    # 简单判断：如果第一行没有数据类型关键字，可能是说明表
    first_row_text = ' '.join(c['text'] for c in t['rows'][0]) if t['rows'] else ''
    has_type = any(kw in all_text for kw in ['UINTEGER', 'UINT', 'bit', 'float', 'INT'])
    has_data = t['row_count'] > 1 and any(c['text'] for row in t['rows'][1:] for c in row if c['text'])
    
    print(f"\n源表#{table_idx+1}: {paras[-1][:60] if paras else '(无标题)'}")
    print(f"  维度: {t['row_count']}行×{t['col_count']}列")
    print(f"  第1行: {[c['text'][:20] for c in t['rows'][0]]}")
    print(f"  含数据类型关键字: {has_type}")
    print(f"  初步角色: {role}")
